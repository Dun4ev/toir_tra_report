from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName

from cmm_builder import generate_comment_sheets


def _create_template(path: Path) -> None:
    """Создаёт упрощённый шаблон CMM с именованными диапазонами."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["D1"] = ""
    ws["D4"] = ""
    ws["D6"] = ""
    wb.defined_names.add(DefinedName(name="ReportName", attr_text="'Sheet1'!$D$1"))
    wb.defined_names.add(DefinedName(name="CreatedDate", attr_text="'Sheet1'!$D$4"))
    wb.defined_names.add(DefinedName(name="ExtraField1", attr_text="'Sheet1'!$D$6"))
    wb.save(path)


def test_generate_comment_sheets_creates_workbook(tmp_path: Path) -> None:
    template_path = tmp_path / "CommentSheet_Template.xltx"
    _create_template(template_path)

    docs_dir = tmp_path / "docs"
    docs_dir.mkdir()
    (docs_dir / "CT-DR-II.7.4-00-1G.docx").write_text("stub", encoding="utf-8")

    tz_map = {"II.7.4": "Описание объекта"}

    result = generate_comment_sheets(
        docs_dir,
        template_path,
        tz_map,
        normalize_key=str.upper,
    )

    assert len(result.created) == 1
    output_path = result.created[0]
    assert output_path.name == "CT-DR-II.7.4-00-1G_CMM.xlsx"
    assert not result.skipped_existing
    assert not result.failed

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["D1"].value == "CT-DR-II.7.4-00-1G"
    assert ws["D6"].value == "Описание объекта"
    wb.close()


def test_generate_comment_sheets_skips_existing(tmp_path: Path) -> None:
    template_path = tmp_path / "CommentSheet_Template.xltx"
    _create_template(template_path)

    docs_dir = tmp_path / "docs"
    docs_dir.mkdir()
    report_path = docs_dir / "CT-DR-II.7.4-00-1G.docx"
    report_path.write_text("stub", encoding="utf-8")

    existing = docs_dir / "CT-DR-II.7.4-00-1G_CMM.xlsx"
    _create_template(existing)

    tz_map: dict[str, str] = {}

    result = generate_comment_sheets(
        docs_dir,
        template_path,
        tz_map,
        normalize_key=str.upper,
    )

    assert not result.created
    assert report_path in result.skipped_existing
    assert not result.failed
