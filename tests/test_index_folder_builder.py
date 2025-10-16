import pytest
from pathlib import Path
from openpyxl import Workbook

from index_folder_builder import prepare_index_folders


def _build_tz_file(
    path: Path, rows: list[tuple[str, str | None, str, str | None]]
) -> None:
    """Создаёт рабочую книгу TZ_glob.xlsx с указанными строками."""
    wb = Workbook()
    ws = wb.active
    ws.title = "gen_cl"

    for idx, (lookup, periodicity, suffix, reserved) in enumerate(rows, start=1):
        ws.cell(row=idx, column=2, value=lookup)
        if periodicity is not None:
            ws.cell(row=idx, column=5, value=periodicity)
        ws.cell(row=idx, column=7, value=suffix)
        if reserved is not None:
            ws.cell(row=idx, column=8, value=reserved)

    wb.save(path)
    wb.close()


def test_prepare_index_folders_groups_files_and_moves(tmp_path: Path) -> None:
    source_dir = tmp_path / "source"
    destination_dir = tmp_path / "dest"
    source_dir.mkdir()
    destination_dir.mkdir()

    tz_file = tmp_path / "TZ_glob.xlsx"
    _build_tz_file(tz_file, [("II.7.4", "1Г", "GST", "00")])

    files = [
        "CT-AAA-TRA-II.7.4-00-1G-20250101-00.pdf",
        "CT-AAA-TRA-II.7.4-00-1G-20250101-00.docx",
        "CT-AAA-TRA-II.2.6-00-C-20250102-00.pdf",
    ]
    for name in files:
        (source_dir / name).write_text("demo", encoding="utf-8")

    status_messages: list[str] = []
    created_dirs = prepare_index_folders(
        source_dir,
        destination_dir,
        tz_file,
        status_callback=status_messages.append,
        use_copy=False,
    )

    group_dir = destination_dir / "II.7.4-00-1G_GST"
    c_dir = destination_dir / "II.2.6-00-C"

    assert group_dir.exists() and c_dir.exists()
    assert sorted(p.name for p in group_dir.iterdir()) == [
        "CT-AAA-TRA-II.7.4-00-1G-20250101-00.docx",
        "CT-AAA-TRA-II.7.4-00-1G-20250101-00.pdf",
    ]
    assert sorted(p.name for p in c_dir.iterdir()) == [
        "CT-AAA-TRA-II.2.6-00-C-20250102-00.pdf",
    ]
    assert not any(source_dir.iterdir())
    assert status_messages and status_messages[-1] == "Группировка завершена."
    assert len(created_dirs) == 2


def test_prepare_index_folders_raises_when_no_matching_files(tmp_path: Path) -> None:
    source_dir = tmp_path / "source"
    destination_dir = tmp_path / "dest"
    source_dir.mkdir()
    destination_dir.mkdir()

    tz_file = tmp_path / "TZ_glob.xlsx"
    _build_tz_file(tz_file, [("II.7.4", "1Г", "GST", "00")])

    (source_dir / "irrelevant.txt").write_text("noop", encoding="utf-8")

    with pytest.raises(ValueError):
        prepare_index_folders(source_dir, destination_dir, tz_file)
