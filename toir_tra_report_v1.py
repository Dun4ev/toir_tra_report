import re
from pathlib import Path
from datetime import datetime
import sys
import os
import subprocess
import zipfile
import webbrowser
import json
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from index_folder_builder import prepare_index_folders
from cmm_builder import generate_comment_sheets

# Настройка UTF-8 вывода
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    messagebox.showerror("Ошибка зависимостей", "Библиотека openpyxl не найдена. Установите ее: pip install openpyxl")
    sys.exit(1)

# ============= НАСТРОЙКИ =============
# Карта статусов и соответствующих им папок
TEMPLATE_STATUSES = {
    "izdato na pregled_GST  (для отправки Заказчику)": "izdato_na_pregled_gst",
    "na uvid_app  (для отправки подписанных отчетов Субподрядчику)": "na_uvid_app",
    "za upotrebu_cmm  (для отправки комментариев Субподрядчику)": "za_upotrebu_cmm",
}
DEFAULT_COMPANY_NAMES = {
    "GST": "Gastrans",
    "CDT": "Comita DTech",
    "CNE": "Contex",
    "GGC": "Giprogazcentr",
    "DGT": "Drager",
    "DTA": "DTA Process&Safety",
    "ENL": "Energo Lab",
    "ENK": "Energointeh Kibernetika",
    "ERG": "ENREGRO",
    "IDP": "IvDam",
    "KBV": "KBV",
    "KNT": "Kontron",
    "KSR": "KSR Service",
    "MRS": "Martin",
    "MWT": "Milanovic",
    "MSV": "MOS-AV",
    "NTK": "Netiks",
    "OST": "Ostral",
    "PCM": "CTO ProChrom",
    "PSI": "Petrolsoft",
    "PTD": "Premi Trade",
    "PMG": "PROMONT GROUP",
    "RIM": "Real Impeks",
    "SBT": "SBT",
    "SNX": "SENERMAX",
    "SIM": "SIEMENS ENERGY",
    "TTP": "TehnoTerm",
    "TMG": "TERMOGAMA",
    "TER": "Termoingenjering",
    "TSL": "Tesla Ekspo",
    "VIS": "VIS Company",
    "VLK": "Vulkan Ingenjering",
}
DEFAULT_TRANSMITTAL_SENDERS = [
    "Андреj Дунаев, главни експерт +381 69 801 53 43, dunaevaa@cmtech.rs",
    "Анатолиј Крамар, главни експерт +381 69 807 65 37, kramaras@cmtech.rs"
]
# --- Определение путей для .exe и обычного режима ---
def get_base_path() -> Path:
    """Возвращает базовый путь для ресурсов, работающий и для .exe."""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        # Запущено из PyInstaller bundle
        return Path(sys.executable).parent
    else:
        # Запущено как обычный .py скрипт
        return Path(__file__).parent

BASE_DIR = get_base_path()
SETTINGS_FILE = BASE_DIR / "settings.json"

# --- Функции для работы с настройками ---
def build_default_settings() -> dict[str, object]:
    """Формирует структуру настроек со значениями по умолчанию."""
    return {
        "templates_path": "",
        "company_names": DEFAULT_COMPANY_NAMES.copy(),
        "senders": list(DEFAULT_TRANSMITTAL_SENDERS),
    }

def save_settings(settings_data: dict):
    """Сохраняет данные в settings.json."""
    try:
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(settings_data, f, indent=2)
        return True
    except Exception as e:
        print(f"[ОШИБКА] Не удалось сохранить settings.json: {e}")
        return False

def load_settings() -> tuple[Path, dict[str, str], list[str]]:
    """Загружает настройки из settings.json или возвращает значения по умолчанию."""
    default_path = BASE_DIR
    default_companies = DEFAULT_COMPANY_NAMES.copy()
    default_senders = list(DEFAULT_TRANSMITTAL_SENDERS)
    
    if not SETTINGS_FILE.exists():
        # Первый запуск: создаем settings.json
        print(f"Файл настроек не найден. Создание нового: {SETTINGS_FILE}")
        save_settings(build_default_settings())
        return default_path, DEFAULT_COMPANY_NAMES.copy(), list(DEFAULT_TRANSMITTAL_SENDERS)

    try:
        with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
            settings = json.load(f)
        
        custom_path_str = settings.get("templates_path")
        final_path = default_path
        if custom_path_str and Path(custom_path_str).is_dir():
            print(f"Используется кастомный путь для шаблонов: {custom_path_str}")
            final_path = Path(custom_path_str)
        else:
            print(f"Кастомный путь не задан или не найден. Используется путь по умолчанию: {default_path}")

        company_names_raw = settings.get("company_names", default_companies)
        if isinstance(company_names_raw, dict):
            company_names = company_names_raw
        else:
            company_names = default_companies

        senders_raw = settings.get("senders", default_senders)
        if isinstance(senders_raw, list) and all(isinstance(item, str) for item in senders_raw):
            senders = senders_raw or default_senders
        else:
            senders = default_senders

        # При необходимости дополняем недостающие поля и сохраняем
        needs_update = False
        if "senders" not in settings:
            settings["senders"] = senders
            needs_update = True
        if "company_names" not in settings:
            settings["company_names"] = company_names
            needs_update = True
        if needs_update:
            save_settings(settings)

        return final_path, company_names, senders

    except (json.JSONDecodeError, Exception) as e:
        print(f"[ОШИБКА] Не удалось прочитать settings.json: {e}")
        return default_path, default_companies, default_senders

def ensure_template_structure(base_path: Path):
    """
    Проверяет и создает необходимую структуру папок для шаблонов.
    Если запущено из .exe и папки Template нет, копирует ее из бандла.
    """
    persistent_template_dir = base_path / "Template"
    print(f"Проверка структуры папок в: {persistent_template_dir}")

    if not persistent_template_dir.exists():
        print(f"Папка 'Template' не найдена в {base_path}. Попытка копирования...")
        
        # Определяем путь к данным внутри бандла
        bundle_dir = Path(sys._MEIPASS) if getattr(sys, 'frozen', False) else Path(__file__).parent
        source_template_dir = bundle_dir / "Template"

        if source_template_dir.exists():
            try:
                shutil.copytree(source_template_dir, persistent_template_dir)
                print(f"Папка 'Template' успешно скопирована из бандла в {persistent_template_dir}")
            except Exception as e:
                print(f"[КРИТИЧЕСКАЯ ОШИБКА] Не удалось скопировать папку 'Template': {e}")
                messagebox.showerror("Критическая ошибка", f"Не удалось создать папку с шаблонами: {e}")
                return
        else:
            print("[ПРЕДУПРЕЖДЕНИЕ] Исходная папка 'Template' не найдена даже в бандле.")

    # В любом случае, убедимся, что подпапки статусов существуют
    try:
        for status_folder in TEMPLATE_STATUSES.values():
            (persistent_template_dir / "template_tra" / status_folder).mkdir(parents=True, exist_ok=True)
        print("Структура папок в порядке.")
    except Exception as e:
        print(f"[ОШИБКА] Не удалось создать подпапки статусов: {e}")

# --- Основные пути ---
TEMPLATES_ROOT, COMPANY_NAMES, TRANSMITTAL_SENDERS = load_settings()
ensure_template_structure(TEMPLATES_ROOT)

TEMPLATE_DIR = TEMPLATES_ROOT / "Template" / "template_tra"
TZ_FILE_PATH = TEMPLATES_ROOT / "Template" / "TZ_glob.xlsx"
COMMENT_TEMPLATE_PATH = TEMPLATES_ROOT / "Template" / "CommentSheet_Template.xltx"

# --- Настройки ячеек и колонок (можно вынести в конфиг) ---
DATE_CELL_ADDR = "C3"
DATE_FMT_TEXT = "%d.%m.%Y"
FOOTER_ANCHOR_NAME = "FooterAnchor"
FIRST_DATA_ROW = 18
COL_RB = 2
COL_BD = 3
COL_NZ = 9
MERGE_BD_FROM, MERGE_BD_TO = 3, 8
MERGE_NZ_FROM, MERGE_NZ_TO = 9, 12
ALLOWED_EXT = { ".pdf", ".docx", ".xlsx", ".xls", ".dwg", ".zip", ".7z"}

# --- Регулярные выражения ---
RE_INDEX = re.compile(
    r"\b([IVXLCDM]+)\.(\d+)(?:\.(\d+))?(?:\.(\d+))?([A-Za-zА-Яа-я])?\b",
    re.IGNORECASE
)
DATE_PATTERNS = [
    re.compile(r"\b\d{2}\.\d{2}\.\d{4}\b"),
    re.compile(r"\b\d{4}-\d{2}-\d{2}\b"),
    re.compile(r"\b\d{2}\.\d{2}\.\d{2}\b"),
]

# ---------- БИЗНЕС-ЛОГИКА (ОСНОВНОЙ КОД ОБРАБОТКИ) ----------

def process_files(
    target_dir: Path,
    template_path: Path,
    status_callback,
    create_archive_flag: bool,
    delete_files_flag: bool,
    sender_value: str | None = None,
):
    """Основная функция для обработки файлов и создания отчета."""
    try:
        status_callback(f"Загрузка шаблона: {template_path.name}")
        if not template_path.exists():
            raise FileNotFoundError(f"Шаблон не найден: {template_path}")
        if not target_dir.exists():
            raise FileNotFoundError(f"Папка с файлами не найдена: {target_dir}")
        if not TZ_FILE_PATH.exists():
            status_callback(f"[ПРЕДУПРЕЖДЕНИЕ] Не найден {TZ_FILE_PATH} — 'Назив документа' будет пустым.")

        wb = load_workbook(template_path)
        ws = wb.active

        if sender_value:
            status_callback("Заполнение данных отправителя...")
            if not set_named_cell_value(wb, "pripmem", sender_value):
                ws["I22"].value = sender_value

        status_callback("Запись даты...")
        write_date(ws)

        footer_row = get_footer_row_by_name(wb, ws.title, FOOTER_ANCHOR_NAME) or 20
        status_callback(f"Найдена строка футера: {footer_row}")

        status_callback(f"Поиск документов в {target_dir}...")
        files = list_docs(target_dir)
        if not files:
            messagebox.showwarning("Нет файлов", f"В папке {target_dir} не найдено файлов для обработки.")
            return

        status_callback(f"Найдено {len(files)} файлов. Чтение карты индексов...")
        tz_map = build_tz_map_from_xlsx(TZ_FILE_PATH)

        num_files = len(files)
        available_data_rows = footer_row - FIRST_DATA_ROW
        rows_to_insert = 0
        if num_files > available_data_rows:
            rows_to_insert = num_files - available_data_rows

        if rows_to_insert > 0:
            status_callback(f"Вставка {rows_to_insert} строк...")
            insert_rows_and_preserve_footer_merges(ws, footer_row, rows_to_insert)

        new_footer_row = footer_row + rows_to_insert
        status_callback("Заполнение строк данными...")
        final_footer_row = fill_rows(ws, files, tz_map, FIRST_DATA_ROW, new_footer_row)
        
        status_callback("Обновление якоря футера и области печати...")
        update_footer_anchor(wb, ws.title, FOOTER_ANCHOR_NAME, final_footer_row)
        
        last_row = ws.max_row
        ws.print_area = f'B3:P{last_row}'

        wb.template = False
        
        prefix = template_path.stem.replace("-Template", "-")
        saved_path = save_with_increment(wb, target_dir, prefix=prefix)
        
        if create_archive_flag:
            status_callback("Создание ZIP-архива...")
            archive_name = saved_path.with_suffix('').name + "_att.zip"
            archive_path = saved_path.parent / archive_name
            
            try:
                with zipfile.ZipFile(archive_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file_to_add in files:
                        zipf.write(file_to_add, arcname=file_to_add.name)
                
                if delete_files_flag:
                    status_callback("Удаление исходных файлов...")
                    try:
                        for file_to_delete in files:
                            os.remove(file_to_delete)
                        status_callback("Исходные файлы удалены. Открываю папку...")
                    except Exception as e:
                        messagebox.showerror("Ошибка удаления", f"Не удалось удалить исходные файлы: {e}")
                        status_callback("Ошибка удаления файлов.")
                else:
                    status_callback("Архив создан. Открываю папку...")

            except Exception as e:
                status_callback(f"Ошибка создания архива: {e}")
                messagebox.showerror("Ошибка архивации", f"Не удалось создать ZIP-архив: {e}")
        else:
             status_callback("Готово! Файл сохранен. Открываю папку...")

        try:
            if sys.platform == "win32":
                os.startfile(saved_path.parent)
            elif sys.platform == "darwin":
                subprocess.run(['open', str(saved_path.parent)])
            else:
                subprocess.run(['xdg-open', str(saved_path.parent)])
        except Exception as e:
            messagebox.showwarning("Ошибка", f"Не удалось автоматически открыть папку: {e}")

    except Exception as e:
        status_callback(f"Ошибка: {e}")
        messagebox.showerror("Ошибка выполнения", f"Произошла ошибка:\n{e}")

# ---------- Утилиты (без изменений) ----------

def list_docs(doc_dir: Path):
    return [p for p in sorted(doc_dir.rglob('*'))
            if p.is_file() and p.suffix.lower() in ALLOWED_EXT]

def write_date(ws):
    today = datetime.now().strftime(DATE_FMT_TEXT)
    cell = ws[DATE_CELL_ADDR]
    val = cell.value
    if isinstance(val, str):
        new = val
        for pat in DATE_PATTERNS:
            if pat.search(new):
                new = pat.sub(today, new, count=1)
                break
        else:
            new = today
        cell.value = new
    else:
        cell.value = today

def set_named_cell_value(workbook, defined_name: str, value: str) -> bool:
    """Записывает значение в именованную ячейку, если она определена."""
    dn = workbook.defined_names.get(defined_name)
    if dn is None:
        return False
    try:
        destinations = list(dn.destinations)
    except Exception:
        destinations = []
    for sheet_name, ref in destinations:
        sheet_key = sheet_name.strip("'") if isinstance(sheet_name, str) else sheet_name
        if sheet_key not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_key]
        coord = str(ref).split("!")[-1].replace("$", "")
        if ":" in coord:
            try:
                cells = worksheet[coord]
            except Exception:
                continue
            if isinstance(cells, tuple):
                first_row = cells[0]
                if not first_row:
                    continue
                target_cell = first_row[0]
            else:
                target_cell = cells
        else:
            try:
                target_cell = worksheet[coord]
            except Exception:
                continue
        target_cell.value = value
        return True
    return False

def normalize_key(key: str) -> str:
    key = key.upper()
    replacements = {'A': 'А', 'B': 'Б', 'V': 'В', 'G': 'Г'}
    for lat, cyr in replacements.items():
        key = key.replace(lat, cyr)
    return key

def get_footer_row_by_name(wb, ws_name: str, name: str) -> int | None:
    dn = wb.defined_names.get(name)
    if dn is None: return None
    try:
        destinations = list(dn.destinations)
    except Exception:
        destinations = []
    for sname, ref in destinations:
        s_clean = sname.strip("'" ) if isinstance(sname, str) else sname
        if s_clean == ws_name:
            coord = str(ref).split("!")[-1].replace("$", "")
            m = re.search(r"\d+", coord)
            if m: return int(m.group(0))
    return None

def update_footer_anchor(wb, ws_name: str, name: str, new_row: int, column_letter: str = "B"):
    ref = f"'{ws_name}'!${column_letter}${new_row}"
    try:
        wb.defined_names.delete(name)
    except Exception:
        pass
    dn_obj = DefinedName(name=name, attr_text=ref)
    try:
        wb.defined_names[name] = dn_obj
    except Exception:
        wb.defined_names.append(dn_obj)

def ensure_row_merges(ws, row, footer_row):
    target_cols_min = min(MERGE_BD_FROM, MERGE_NZ_FROM)
    target_cols_max = max(MERGE_BD_TO, MERGE_NZ_TO)
    to_unmerge = []
    for mr in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = mr.bounds
        if max_row >= footer_row: continue
        if (min_row <= row <= max_row) and not (max_col < target_cols_min or min_col > target_cols_max):
            to_unmerge.append(str(mr))
    for ref in to_unmerge:
        try:
            ws.unmerge_cells(ref)
        except Exception:
            pass
    rng1 = f"{get_column_letter(MERGE_BD_FROM)}{row}:{get_column_letter(MERGE_BD_TO)}{row}"
    rng2 = f"{get_column_letter(MERGE_NZ_FROM)}{row}:{get_column_letter(MERGE_NZ_TO)}{row}"
    ws.merge_cells(rng1)
    ws.merge_cells(rng2)

def build_tz_map_from_xlsx(xlsx_path: Path) -> dict[str, str]:
    tz_map: dict[str, str] = {}
    if not xlsx_path.exists(): return tz_map
    wb = load_workbook(xlsx_path, data_only=True)
    for ws in wb.worksheets:
        max_col = min(ws.max_column, 20)
        for r in range(1, ws.max_row + 1):
            idx_val, idx_col = None, None
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    m = RE_INDEX.search(v)
                    if m:
                        roman, num1, num2, num3, suf = m.groups()
                        suf = suf or ""
                        idx_val = f"{roman.upper()}.{num1}"
                        if num2: idx_val += f".{num2}"
                        if num3: idx_val += f".{num3}"
                        idx_val += suf
                        idx_col = c
                        break
            if not idx_val: continue
            naziv = None
            vC = ws.cell(r, 3).value
            if isinstance(vC, str) and vC.strip():
                naziv = vC.strip()
            else:
                for c in range((idx_col or 1) + 1, max_col + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str) and len(v.strip()) >= 3:
                        naziv = v.strip()
                        break
            if naziv:
                normalized_key = normalize_key(idx_val)
                if normalized_key not in tz_map:
                    tz_map[normalized_key] = naziv
    return tz_map

def extract_index_from_name(filename: str) -> str | None:
    m = RE_INDEX.search(filename)
    if not m: return None
    roman, num1, num2, num3, suf = m.groups()
    suf = suf or ""
    idx = f"{roman.upper()}.{num1}"
    if num2: idx += f".{num2}"
    if num3: idx += f".{num3}"
    idx += suf
    return idx

def insert_rows_and_preserve_footer_merges(ws, insert_at_row: int, num_rows: int):
    if num_rows <= 0: return
    MAX_COL_TO_COPY = 20
    footer_start_row = insert_at_row
    footer_end_row = ws.max_row
    if footer_end_row < footer_start_row:
        ws.insert_rows(insert_at_row, amount=num_rows)
        return
    footer_snapshot = []
    for r_idx in range(footer_start_row, footer_end_row + 1):
        row_dim = ws.row_dimensions[r_idx]
        row_info = {"height": row_dim.height, "cells": []}
        for c_idx in range(1, MAX_COL_TO_COPY + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            row_info["cells"].append((cell.value, cell._style))
        footer_snapshot.append(row_info)
    footer_merges = [mr for mr in list(ws.merged_cells.ranges) if mr.min_row >= footer_start_row]
    for mr in footer_merges:
        ws.unmerge_cells(str(mr))
    ws.insert_rows(insert_at_row, amount=num_rows)
    new_footer_start_row = footer_start_row + num_rows
    for r_offset, row_info in enumerate(footer_snapshot):
        new_row_num = new_footer_start_row + r_offset
        if row_info["height"] is not None:
            ws.row_dimensions[new_row_num].height = row_info["height"]
        for c_offset, (value, style) in enumerate(row_info["cells"]):
            col_num = 1 + c_offset
            new_cell = ws.cell(row=new_row_num, column=col_num)
            new_cell.value = value
            new_cell._style = style
    for mr in footer_merges:
        mr.shift(0, num_rows)
        ws.merge_cells(str(mr))

def fill_rows(ws, files, tz_map: dict, start_row: int, final_footer_row: int):
    min_col_style, max_col_style = 2, 16
    template_styles = [ws.cell(row=start_row, column=j)._style for j in range(min_col_style, max_col_style + 1)]
    template_row_height = ws.row_dimensions[start_row].height
    const_vals = {
        13: ws.cell(row=start_row, column=13).value,
        14: ws.cell(row=start_row, column=14).value,
        15: ws.cell(row=start_row, column=15).value,
    }
    for i, p in enumerate(files, 1):
        r = start_row + i - 1
        if r >= final_footer_row: continue
        if r > start_row:
            if template_row_height is not None:
                ws.row_dimensions[r].height = template_row_height
            for j_idx, style in enumerate(template_styles):
                ws.cell(row=r, column=min_col_style + j_idx)._style = style
        ensure_row_merges(ws, r, final_footer_row)
        ws.cell(r, COL_RB).value = i
        c = ws.cell(r, COL_BD)
        c.value = p.name
        c.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        idx = extract_index_from_name(p.name)
        base_naziv = tz_map.get(normalize_key(idx), "") if idx else ""
        final_naziv = ""
        if base_naziv:
            prefix = ""
            if "-C-" in p.name.upper(): prefix += "Корективно одржавање. "
            if "_CMM" in p.name.upper(): prefix += "Листа коментара уз документ. "
            final_naziv = prefix + base_naziv
        naziv_cell = ws.cell(r, COL_NZ)
        naziv_cell.value = final_naziv
        naziv_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        for col_num, value in const_vals.items():
            ws.cell(row=r, column=col_num).value = value
    return final_footer_row

def save_with_increment(wb, out_dir: Path, prefix="CT-GST-TRA-PRM-"):
    out_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%y%m%d")
    n = 1
    while True:
        out = out_dir / f"{prefix}{today}_{n:02d}.xlsx"
        if not out.exists():
            wb.save(out)
            return out
        n += 1

# ---------- ГРАФИЧЕСКИЙ ИНТЕРФЕЙС (GUI) ----------

def create_transmittal_gui():
    """Создает и управляет GUI для выбора папки и шаблона."""
    root = tk.Tk()
    root.title("Формирование трансмиттала v3.0")
    root.geometry("550x790")
    root.resizable(False, False)

    # --- Стилизация ---
    BG_COLOR = "#F4F6F5"
    FRAME_COLOR = "#FFFFFF"
    ACTIVE_CARD_COLOR = "#E3F2FD"
    BUTTON_COLOR = "#4CAF50"
    BUTTON_ACTIVE_COLOR = "#45a049"
    TEXT_COLOR = "#333333"
    DISABLED_TEXT_COLOR = "#aaaaaa"
    STATUS_BAR_COLOR = "#E0E0E0"
    FONT_NORMAL = ("Segoe UI", 10)
    FONT_BOLD = ("Segoe UI", 11, "bold")
    FONT_LABEL = ("Segoe UI", 9)
    FONT_HELP_TEXT = ("Segoe UI", 8)

    root.config(bg=BG_COLOR)

    style = ttk.Style(root)
    style.theme_use('clam')

    style.configure("TButton", background=BUTTON_COLOR, foreground="white", font=FONT_BOLD, bordercolor=BUTTON_COLOR, lightcolor=BUTTON_COLOR, darkcolor=BUTTON_COLOR, padding=(10, 8))
    style.map("TButton", background=[('active', BUTTON_ACTIVE_COLOR)], foreground=[('active', 'white')])
    style.configure("TMenubutton", background="white", foreground=TEXT_COLOR, font=FONT_NORMAL, arrowcolor=TEXT_COLOR, bordercolor=STATUS_BAR_COLOR)
    style.configure("TFrame", background=BG_COLOR)
    style.configure("TLabel", background=BG_COLOR, foreground=TEXT_COLOR, font=FONT_NORMAL)
    style.configure("Header.TLabel", font=FONT_BOLD, background=FRAME_COLOR)
    style.configure("Status.TLabel", background=STATUS_BAR_COLOR, foreground=TEXT_COLOR, padding=5, font=("Segoe UI", 9))
    style.configure("Card.TFrame", background=FRAME_COLOR)
    style.configure("ActiveCard.TFrame", background=ACTIVE_CARD_COLOR)
    style.configure("TCheckbutton", background=FRAME_COLOR, font=FONT_NORMAL, foreground=TEXT_COLOR)
    style.map("TCheckbutton", foreground=[('disabled', DISABLED_TEXT_COLOR)])
    style.configure("TRadiobutton", background=FRAME_COLOR, font=FONT_NORMAL, foreground=TEXT_COLOR)
    style.map("TRadiobutton", background=[('active', BG_COLOR)])


    notebook = ttk.Notebook(root)
    notebook.pack(fill=tk.BOTH, expand=True)

    report_tab = ttk.Frame(notebook, padding=0)
    index_tab = ttk.Frame(notebook, padding=0)
    cmm_tab = ttk.Frame(notebook, padding=0)
    notebook.add(report_tab, text="Формирование трансмиттала")
    notebook.add(index_tab, text="Формирование папок")
    notebook.add(cmm_tab, text="Создание CMM")


    # --- Переменные ---
    selected_folder = tk.StringVar()
    selected_status_key = tk.StringVar(value=list(TEMPLATE_STATUSES.keys())[0])
    selected_template_key = tk.StringVar()
    selected_sender = tk.StringVar(value=TRANSMITTAL_SENDERS[0] if TRANSMITTAL_SENDERS else "")
    should_create_archive = tk.BooleanVar(value=True)
    should_delete_files = tk.BooleanVar(value=False)
    
    templates_map = {}
    index_source_path = tk.StringVar()
    index_destination_path = tk.StringVar()
    index_source_display = tk.StringVar(value="(не выбрана)")
    index_destination_display = tk.StringVar(value="(не выбрана)")
    index_status_message = tk.StringVar(value="Выберите исходную и целевую папки.")
    should_copy_files = tk.BooleanVar(value=True)
    should_group_by_suffix = tk.BooleanVar(value=True)
    cmm_source_path = tk.StringVar()
    cmm_source_display = tk.StringVar(value='(Не выбрано)')
    cmm_status_message = tk.StringVar(value='Готово к запуску.')
    cmm_run_button: ttk.Button | None = None

    # --- Функции-обработчики GUI ---
    def select_custom_template_path():
        """Диалог выбора и сохранения нового пути к папке с шаблонами."""
        folder_path = filedialog.askdirectory(title="Выберите корневую папку с вашими шаблонами (внутри нее должна быть папка Template)")
        if not folder_path:
            return

        # Читаем текущие настройки, чтобы не затереть другие возможные параметры
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                settings = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            settings = {}
        
        settings["templates_path"] = folder_path
        if save_settings(settings):
            messagebox.showinfo(
                "Настройка сохранена",
                f"Новый путь к шаблонам сохранен.\n\n{folder_path}\n\nПожалуйста, перезапустите программу, чтобы применить изменения."
            )
        else:
            messagebox.showerror("Ошибка", "Не удалось сохранить файл настроек.")


    def open_github(event=None):
        webbrowser.open_new("https://github.com/Dun4ev/toir_tra_report")

    def update_template_options(*args):
        nonlocal templates_map
        status_dir_name = TEMPLATE_STATUSES.get(selected_status_key.get())
        if not status_dir_name:
            return

        templates_path = TEMPLATE_DIR / status_dir_name
        templates_map.clear()

        if templates_path.is_dir():
            for f in templates_path.glob("*.xltx"):
                # Пример: CT-GST-TRA-PRM-Template.xltx -> GST
                parts = f.stem.split('-')
                if len(parts) > 1:
                    abbr = parts[1].upper()
                    # Ищем полное имя в настройках, если нет - используем саму аббревиатуру
                    full_name = COMPANY_NAMES.get(abbr, abbr)
                    
                    # Формируем ключ для отображения в меню
                    full_name = COMPANY_NAMES.get(abbr, abbr)
                    if abbr == "XXX":
                        full_name = COMPANY_NAMES.get("XXX", "Общий")
                    
                    key_name = f"({abbr}) {full_name}"

                    templates_map[key_name] = f.name
        
        # Обновление меню шаблонов
        menu = template_menu["menu"]
        menu.delete(0, "end")
        
        if not templates_map:
            template_menu.config(state=tk.DISABLED)
            selected_template_key.set("")
            return
        
        template_menu.config(state=tk.NORMAL)
        # Сортируем ключи, чтобы "Общий" был в конце
        sorted_keys = sorted(templates_map.keys(), key=lambda x: "zzz" if "Общий" in x else x)
        for key in sorted_keys:
            menu.add_command(label=key, command=tk._setit(selected_template_key, key))
        
        # Автоматический выбор шаблона
        folder_path = selected_folder.get()
        default_key = COMPANY_NAMES.get("XXX", "Общий")

        if folder_path:
            folder_name_upper = Path(folder_path).name.upper()
            # --- НОВАЯ ЛОГИКА: Динамическое определение аббревиатур ---
            # 1. Извлекаем аббревиатуры из имен найденных шаблонов
            available_abbrs = []
            for template_filename in templates_map.values():
                # Пример: CT-GST-TRA-PRM-Template.xltx -> ['CT', 'GST', 'TRA', 'PRM', 'Template.xltx']
                parts = template_filename.split('-')
                if len(parts) > 1 and parts[1].upper() != "XXX":
                    available_abbrs.append(parts[1].upper())
            
            # 2. Ищем совпадение в имени папки
            found_template = False
            # Сортируем по длине, чтобы сначала проверять более длинные и специфичные аббревиатуры
            for abbr in sorted(available_abbrs, key=len, reverse=True):
                if f"_{abbr}" in folder_name_upper or f"-{abbr}" in folder_name_upper or re.search(rf'\b{abbr}\b', folder_name_upper):
                    # Нашли аббревиатуру, теперь найдем ключ шаблона (его отображаемое имя), которому она принадлежит
                    for key, filename in templates_map.items():
                        if f"-{abbr}-" in filename.upper():
                            selected_template_key.set(key)
                            found_template = True
                            break
                if found_template:
                    break
            
            # 3. Если ничего не найдено, используем шаблон по умолчанию
            if not found_template:
                if default_key in templates_map:
                    selected_template_key.set(default_key)
                # Доп. логика: если не нашли по аббревиатуре и default_key отсутствует,
                # пытаемся выбрать (XXX) Общий среди доступных шаблонов
                if not found_template:
                    xxx_key = None
                    for k, fname in templates_map.items():
                        if "-XXX-" in str(fname).upper():
                            xxx_key = k
                            break
                    if xxx_key is None:
                        for k in templates_map.keys():
                            if "(XXX)" in str(k).upper():
                                xxx_key = k
                                break
                    if xxx_key:
                        selected_template_key.set(xxx_key)
        else:
            if default_key in templates_map:
                selected_template_key.set(default_key)
            else:
                # Если нет даже папки, выбираем первый в списке
                selected_template_key.set(sorted_keys[0] if sorted_keys else "")
        # Дополнительный запасной вариант: выбрать XXX, если по-прежнему ничего не выбрано
        if not selected_template_key.get():
            for key, filename in templates_map.items():
                fn_upper = str(filename).upper()
                if "-XXX-" in fn_upper or "(XXX)" in str(key).upper():
                    selected_template_key.set(key)
                    break

    def toggle_delete_option():
        if should_create_archive.get():
            delete_check.config(state=tk.NORMAL)
        else:
            delete_check.config(state=tk.DISABLED)
            should_delete_files.set(False)

    def select_folder():
        folder_path = filedialog.askdirectory(title="Выберите папку с документами")
        if folder_path:
            selected_folder.set(folder_path)
            folder_display_label.config(text=f"...{folder_path[-50:]}")
            # --- Обновление ссылки на папку ---
            folder_link_label.config(text=f"🔗 {Path(folder_path).name}")
            update_template_options()

    def run_processing():
        target_dir = selected_folder.get()
        if not target_dir:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите папку с документами.")
            return
        
        sender_value = selected_sender.get().strip()
        if not sender_value:
            messagebox.showerror("Ошибка", "Пожалуйста, выберите отправителя трансмиттала.")
            return

        status_dir_name = TEMPLATE_STATUSES.get(selected_status_key.get())
        template_file_name = templates_map.get(selected_template_key.get())

        if not status_dir_name or not template_file_name:
            messagebox.showerror("Ошибка", "Не удалось определить путь к шаблону. Проверьте выбор статуса и шаблона.")
            return

        template_path = TEMPLATE_DIR / status_dir_name / template_file_name

        run_button.config(state=tk.DISABLED)
        def status_update(message):
            status_label.config(text=message)
            root.update_idletasks()

        process_files(
            Path(target_dir),
            template_path,
            status_update,
            should_create_archive.get(),
            should_delete_files.get(),
            sender_value,
        )
        run_button.config(state=tk.NORMAL)

    # --- Компоновка ---
    main_frame = ttk.Frame(report_tab, padding=(15, 10))
    main_frame.pack(fill=tk.BOTH, expand=True)

    # Блок 1: Выбор папки
    folder_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    folder_card.pack(fill=tk.X, pady=5)
    ttk.Label(folder_card, text="1. Выберите папку с документами", style="Header.TLabel").pack(anchor="w")
    folder_display_label = ttk.Label(folder_card, text="(не выбрана)", font=FONT_LABEL, foreground="#757575", background=FRAME_COLOR)
    folder_display_label.pack(anchor="w", pady=(5, 10))
    ttk.Button(folder_card, text="Выбрать папку...", command=select_folder, style="TButton").pack(anchor="w")

    # Блок 2: Выбор отправителя
    sender_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    sender_card.pack(fill=tk.X, pady=5)
    ttk.Label(sender_card, text="2. Выберите отправителя трансмиттала", style="Header.TLabel").pack(anchor="w")
    ttk.Label(
        sender_card,
        text="Список отправителей настраивается в settings.json.",
        font=FONT_HELP_TEXT,
        foreground="#757575",
        background=FRAME_COLOR,
        justify=tk.LEFT,
    ).pack(anchor="w", pady=(5, 10))
    sender_menu = ttk.OptionMenu(sender_card, selected_sender, selected_sender.get(), *TRANSMITTAL_SENDERS)
    sender_menu.pack(fill=tk.X)

    # Блок 3: Выбор статуса отправки
    status_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    status_card.pack(fill=tk.X, pady=5)
    ttk.Label(status_card, text="3. Выберите статус отправки", style="Header.TLabel").pack(anchor="w", pady=(0, 5))
    
    for status_text in TEMPLATE_STATUSES.keys():
        rb = ttk.Radiobutton(status_card, text=status_text, variable=selected_status_key, value=status_text, style="TRadiobutton")
        rb.pack(anchor="w", padx=5)

    # Блок 4: Выбор шаблона
    template_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    template_card.pack(fill=tk.X, pady=5)
    ttk.Label(template_card, text="4. Выберите компанию (шаблон)", style="Header.TLabel").pack(anchor="w")
    
    info_text = ("Подсказка: шаблон выбирается автоматически, если имя папки содержит (GST, TER и т.д.).")
    info_label = ttk.Label(template_card, text=info_text, font=FONT_HELP_TEXT, foreground="#757575", background=FRAME_COLOR, justify=tk.LEFT)
    info_label.pack(anchor='w', pady=(5, 10))

    template_menu = ttk.OptionMenu(template_card, selected_template_key, "", style="TMenubutton")
    template_menu.pack(fill=tk.X)
    template_menu.config(state=tk.DISABLED)

    # Блок 5: Запуск
    run_card = ttk.Frame(main_frame, style="Card.TFrame", padding=15)
    run_card.pack(fill=tk.X, pady=5)
    
    archive_check = ttk.Checkbutton(run_card, text="Создать ZIP-архив с вложениями", variable=should_create_archive, style="TCheckbutton", command=toggle_delete_option)
    archive_check.pack(anchor="w")

    delete_check = ttk.Checkbutton(run_card, text="Удалить исходные файлы после архивации", variable=should_delete_files, style="TCheckbutton")
    delete_check.pack(anchor="w", padx=(20, 0), pady=(0, 15))

    run_button = ttk.Button(run_card, text="Сформировать трансмиттал", command=run_processing, style="TButton")
    run_button.pack(ipady=10, fill=tk.X)

    # --- Верхнее меню ---
    menubar = tk.Menu(root)
    settings_menu = tk.Menu(menubar, tearoff=0)
    settings_menu.add_command(label="Указать папку с шаблонами...", command=select_custom_template_path)
    menubar.add_cascade(label="Настройки", menu=settings_menu)
    root.config(menu=menubar)

    # --- Нижняя панель (статус-бар и ссылка) ---
    bottom_frame = tk.Frame(root, bg=STATUS_BAR_COLOR)
    bottom_frame.pack(side=tk.BOTTOM, fill=tk.X)

    status_label = ttk.Label(bottom_frame, text="Ожидание...", style="Status.TLabel", anchor="w")
    status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

    # --- Ссылка на папку ---
    folder_link_label = tk.Label(
        bottom_frame,
        text="",
        fg="#00529B",
        cursor="hand2",
        bg=STATUS_BAR_COLOR,
        font=("Segoe UI", 9, "underline")
    )
    folder_link_label.pack(side=tk.LEFT, padx=10)

    def open_selected_folder(event=None):
        folder_path = selected_folder.get()
        if folder_path and Path(folder_path).is_dir():
            try:
                if sys.platform == "win32":
                    os.startfile(folder_path)
                elif sys.platform == "darwin":
                    subprocess.run(['open', folder_path])
                else:
                    subprocess.run(['xdg-open', folder_path])
            except Exception as e:
                messagebox.showwarning("Ошибка", f"Не удалось открыть папку: {e}")

    folder_link_label.bind("<Button-1>", open_selected_folder)


    link_label = tk.Label(bottom_frame, text="🔗 GitHub", fg="blue", cursor="hand2", bg=STATUS_BAR_COLOR, font=("Segoe UI", 8, "underline"))
    link_label.pack(side=tk.RIGHT, padx=10)
    link_label.bind("<Button-1>", open_github)

    def _shorten_path_for_display(path: str) -> str:
        if len(path) <= 50:
            return path
        return f"...{path[-50:]}"

    def update_index_status(message: str) -> None:
        index_status_message.set(message)
        status_label.config(text=message)
        root.update_idletasks()

    def select_index_source_folder() -> None:
        folder_path = filedialog.askdirectory(title="Выбор исходной папки")
        if folder_path:
            index_source_path.set(folder_path)
            index_source_display.set(_shorten_path_for_display(folder_path))
            update_index_status("Исходная папка выбрана.")

    def select_index_destination_folder() -> None:
        folder_path = filedialog.askdirectory(title="Выбор целевой папки")
        if folder_path:
            index_destination_path.set(folder_path)
            index_destination_display.set(_shorten_path_for_display(folder_path))
            update_index_status("Целевая папка выбрана.")

    def open_index_destination_folder(event=None):
        folder_path = index_destination_path.get()
        if folder_path and Path(folder_path).is_dir():
            try:
                if sys.platform == "win32":
                    os.startfile(folder_path)
                elif sys.platform == "darwin":
                    subprocess.run(['open', folder_path])
                else:
                    subprocess.run(['xdg-open', folder_path])
            except Exception as e:
                messagebox.showwarning("Ошибка", f"Не удалось открыть папку: {e}")

    def run_index_packaging() -> None:
        source_dir = index_source_path.get()
        destination_dir = index_destination_path.get()

        if not source_dir:
            messagebox.showerror("Ошибка", "Не выбрана исходная папка.")
            return
        if not destination_dir:
            messagebox.showerror("Ошибка", "Не выбрана целевая папка.")
            return
        if not TZ_FILE_PATH.exists():
            messagebox.showerror("Ошибка", f"Файл {TZ_FILE_PATH.name} не найден.")
            return

        apply_index_button.config(state=tk.DISABLED)
        try:
            update_index_status("Запуск группировки...")
            created_dirs = prepare_index_folders(
                Path(source_dir),
                Path(destination_dir),
                TZ_FILE_PATH,
                status_callback=update_index_status,
                use_copy=should_copy_files.get(),
                group_by_suffix=should_group_by_suffix.get(),
            )
        except FileNotFoundError as exc:
            messagebox.showerror("Ошибка", str(exc))
            update_index_status(str(exc))
        except ValueError as exc:
            # Для ошибок валидации (например, отсутствие суффиксов) показываем более детальное сообщение
            messagebox.showerror("Ошибка группировки", str(exc))
            update_index_status(f"Ошибка: {exc}")
        except Exception as exc:
            messagebox.showerror("Ошибка", f"Неожиданная ошибка: {exc}")
            update_index_status("Возникла ошибка при группировке.")
        else:
            summary = f"Готово: создано {len(created_dirs)} папок."
            update_index_status(summary)
            messagebox.showinfo("Готово", summary)

            # Обновляем ссылку в статус-баре
            dest_path = Path(destination_dir)
            folder_link_label.config(text=f"🔗 {dest_path.name}")
            folder_link_label.unbind("<Button-1>")
            folder_link_label.bind("<Button-1>", lambda e: open_index_destination_folder(e))

            # Автоматически открываем папку
            try:
                if sys.platform == "win32":
                    os.startfile(destination_dir)
                elif sys.platform == "darwin":
                    subprocess.run(['open', destination_dir])
                else:
                    subprocess.run(['xdg-open', destination_dir])
            except Exception as e:
                messagebox.showwarning("Ошибка", f"Не удалось автоматически открыть папку: {e}")
        finally:
            apply_index_button.config(state=tk.NORMAL)

    def update_cmm_status(message: str) -> None:
        cmm_status_message.set(message)
        root.update_idletasks()

    def select_cmm_source_folder() -> None:
        folder_path = filedialog.askdirectory(title='Выберите каталог с отчётами (CT-DR-*)')
        if folder_path:
            cmm_source_path.set(folder_path)
            cmm_source_display.set(_shorten_path_for_display(folder_path))
            update_cmm_status('Каталог выбран.')
        else:
            update_cmm_status('Выбор отменён.')

    def run_cmm_generation() -> None:
        nonlocal cmm_run_button
        folder_value = cmm_source_path.get()
        if not folder_value:
            messagebox.showwarning('Внимание', 'Выберите папку с файлами.')
            return
        if not COMMENT_TEMPLATE_PATH.exists():
            messagebox.showerror('Ошибка', f'Шаблон не найден: {COMMENT_TEMPLATE_PATH}')
            return
        if not TZ_FILE_PATH.exists():
            messagebox.showerror('Ошибка', f'Файл {TZ_FILE_PATH.name} не найден.')
            return

        update_cmm_status('Подготовка к созданию CMM...')
        if cmm_run_button is not None:
            cmm_run_button.config(state=tk.DISABLED)

        try:
            tz_map = build_tz_map_from_xlsx(TZ_FILE_PATH)
            result = generate_comment_sheets(
                Path(folder_value),
                COMMENT_TEMPLATE_PATH,
                tz_map,
                normalize_key,
                status_callback=update_cmm_status,
            )
        except FileNotFoundError as exc:
            messagebox.showerror('Ошибка', str(exc))
            update_cmm_status(str(exc))
            return
        except Exception as exc:
            messagebox.showerror('Ошибка', f'Не удалось создать CMM: {exc}')
            update_cmm_status('Ошибка при создании CMM.')
            return
        finally:
            if cmm_run_button is not None:
                cmm_run_button.config(state=tk.NORMAL)

        summary_lines = [
            f'Создано файлов: {len(result.created)}',
            f'Пропущено (CMM уже существует): {len(result.skipped_existing)}',
        ]
        error_preview = ''
        if result.failed:
            summary_lines.append(f'Ошибки: {len(result.failed)}')
            error_preview = '\n'.join(f'- {path.name}: {error}' for path, error in result.failed[:3])
        message = '\n'.join(summary_lines)
        if result.failed:
            messagebox.showwarning('Создание CMM', f"{message}\n\n{error_preview}")
        else:
            messagebox.showinfo('Создание CMM', message)
        update_cmm_status('Готово.')

    cmm_tab_container = ttk.Frame(cmm_tab, padding=(15, 10))
    cmm_tab_container.pack(fill=tk.BOTH, expand=True)

    cmm_folder_card = ttk.Frame(cmm_tab_container, style="Card.TFrame", padding=15)
    cmm_folder_card.pack(fill=tk.X, pady=5)
    ttk.Label(cmm_folder_card, text="1. Выберите папку с документами CT-DR", style="Header.TLabel").pack(anchor="w")
    ttk.Label(
        cmm_folder_card,
        textvariable=cmm_source_display,
        font=FONT_LABEL,
        foreground="#757575",
        background=FRAME_COLOR,
    ).pack(anchor="w", pady=(5, 10))
    ttk.Button(
        cmm_folder_card,
        text="Выбрать папку...",
        command=select_cmm_source_folder,
        style="TButton",
    ).pack(anchor="w")

    cmm_info_card = ttk.Frame(cmm_tab_container, style="Card.TFrame", padding=15)
    cmm_info_card.pack(fill=tk.X, pady=5)
    ttk.Label(
        cmm_info_card,
        text="Шаблон: CommentSheet_Template.xltx",
        font=FONT_HELP_TEXT,
        foreground="#757575",
        background=FRAME_COLOR,
        justify=tk.LEFT,
        wraplength=480,
    ).pack(anchor="w")

    cmm_run_card = ttk.Frame(cmm_tab_container, style="Card.TFrame", padding=15)
    cmm_run_card.pack(fill=tk.X, pady=5)
    cmm_run_button = ttk.Button(
        cmm_run_card,
        text="Создать_CMM",
        command=run_cmm_generation,
        style="TButton",
    )
    cmm_run_button.pack(fill=tk.X, ipady=10)
    ttk.Label(
        cmm_run_card,
        textvariable=cmm_status_message,
        font=FONT_HELP_TEXT,
        foreground="#757575",
        background=FRAME_COLOR,
        justify=tk.LEFT,
        wraplength=480,
    ).pack(anchor="w", pady=(5, 0))

    index_tab_container = ttk.Frame(index_tab, padding=0)
    index_tab_container.pack(fill=tk.BOTH, expand=True)

    # --- Нижняя часть с кнопкой и статусом --- (перемещена ниже блока информации)

    # --- Верхняя, основная часть ---
    main_content_frame = ttk.Frame(index_tab_container, padding=(15, 0))
    main_content_frame.pack(fill=tk.BOTH, expand=False)

    ttk.Label(
        main_content_frame,
        text="Выберите исходную и целевую папки, затем запустите группировку.",
        font=FONT_HELP_TEXT,
        foreground="#757575",
        background=BG_COLOR,
        wraplength=480,
        justify=tk.LEFT,
    ).pack(fill=tk.X, pady=(0, 5))

    source_card = ttk.Frame(main_content_frame, style="Card.TFrame", padding=15)
    source_card.pack(fill=tk.X, pady=5)
    ttk.Label(source_card, text="1. Исходная папка", style="Header.TLabel").pack(anchor="w")
    ttk.Label(
        source_card,
        textvariable=index_source_display,
        font=FONT_LABEL,
        foreground="#757575",
        background=FRAME_COLOR,
    ).pack(anchor="w", pady=(5, 10))
    ttk.Button(
        source_card,
        text="Выбрать исходную...",
        command=select_index_source_folder,
        style="TButton",
    ).pack(anchor="w")

    destination_card = ttk.Frame(main_content_frame, style="Card.TFrame", padding=15)
    destination_card.pack(fill=tk.X, pady=5)
    ttk.Label(destination_card, text="2. Целевая папка", style="Header.TLabel").pack(anchor="w")
    
    dest_link_label = tk.Label(
        destination_card,
        textvariable=index_destination_display,
        font=("Segoe UI", 9, "underline"),
        fg="#00529B",
        cursor="hand2",
        bg=FRAME_COLOR,
    )
    dest_link_label.pack(anchor="w", pady=(5, 10))
    dest_link_label.bind("<Button-1>", open_index_destination_folder)

    ttk.Button(
        destination_card,
        text="Выбрать целевую...",
        command=select_index_destination_folder,
        style="TButton",
    ).pack(anchor="w")

    options_card = ttk.Frame(main_content_frame, style="Card.TFrame", padding=15)
    options_card.pack(fill=tk.X, pady=5)

    copy_check = ttk.Checkbutton(
        options_card,
        text="Копировать файлы (не перемещать)",
        variable=should_copy_files,
        style="TCheckbutton",
    )
    copy_check.pack(anchor="w")

    group_check = ttk.Checkbutton(
        options_card,
        text="Группировать по суффиксам компаний (напр., 'ENK', 'OST')",
        variable=should_group_by_suffix,
        style="TCheckbutton",
    )
    group_check.pack(anchor="w", pady=(0, 10))

    # --- Место для информации пользователя ---
    info_card = ttk.Frame(main_content_frame, padding=5)
    info_card.pack(fill=tk.BOTH, expand=False, pady=0)

    # --- Левая колонка ---
    col1_wrap = tk.Frame(
        info_card,
        bg=FRAME_COLOR,
        highlightthickness=0,
        highlightbackground=FRAME_COLOR,
        highlightcolor=FRAME_COLOR,
        bd=0,
    )
    col1_wrap.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
    col1_frame = ttk.Frame(col1_wrap, style="Card.TFrame")
    col1_frame.pack(fill=tk.BOTH, expand=True)
    col1_header_label = tk.Label(
        col1_frame,
        text="Группировка ВКЛЮЧЕНА:",
        font=("Segoe UI", 9, "bold"),
        bg=FRAME_COLOR,
        fg=TEXT_COLOR,
        anchor="w",
        justify=tk.LEFT,
    )
    col1_header_label.pack(anchor="w", pady=(0, 5))
    group_on_text = (
        "ENK/\n"
        "├─ I.12.4a-00-2M_ENK/\n"
        "├─ I.12.4v-00-5G_ENK/\n"
        "├─ I.12.6b-00-5G_ENK/\n"
        "└─ II.2.4-00-6M_ENK/"
    )
    line_break = "\n"
    col1_body_label = tk.Label(
        col1_frame,
        text=group_on_text.replace("\n", line_break),
        font=("Segoe UI", 8),
        bg=FRAME_COLOR,
        fg=TEXT_COLOR,
        anchor="w",
        justify=tk.LEFT,
    )
    col1_body_label.pack(anchor="w")

    # --- Правая колонка ---
    col2_wrap = tk.Frame(
        info_card,
        bg=FRAME_COLOR,
        highlightthickness=0,
        highlightbackground=FRAME_COLOR,
        highlightcolor=FRAME_COLOR,
        bd=0,
    )
    col2_wrap.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 0))
    col2_frame = ttk.Frame(col2_wrap, style="Card.TFrame")
    col2_frame.pack(fill=tk.BOTH, expand=True)
    col2_header_label = tk.Label(
        col2_frame,
        text="Группировка ВЫКЛЮЧЕНА:",
        font=("Segoe UI", 9, "bold"),
        bg=FRAME_COLOR,
        fg=TEXT_COLOR,
        anchor="w",
        justify=tk.LEFT,
    )
    col2_header_label.pack(anchor="w", pady=(0, 5))
    group_off_text = (
        "//\n"
        "├─ I.7.1-02-1M_KBV/\n"
        "├─ II.2.2-00-6M_ENK_VLK/\n"
        "├─ II.23.2-00-6M_OST/\n"
        "└─ II.23.3-01-6M_OST/"
    )
    col2_body_label = tk.Label(
        col2_frame,
        text=group_off_text.replace("\n", line_break),
        font=("Segoe UI", 8),
        bg=FRAME_COLOR,
        fg=TEXT_COLOR,
        anchor="w",
        justify=tk.LEFT,
    )
    col2_body_label.pack(anchor="w")
    
    # --- Кликабельность колонок для переключения группировки ---
    def _set_grouping_enabled(event=None):
        should_group_by_suffix.set(True)

    def _set_grouping_disabled(event=None):
        should_group_by_suffix.set(False)

    # Привязки клика по левой колонке (включить)
    for widget in (col1_wrap, col1_frame, col1_header_label, col1_body_label):
        widget.bind("<Button-1>", _set_grouping_enabled)
    # Привязки клика по правой колонке (выключить)
    for widget in (col2_wrap, col2_frame, col2_header_label, col2_body_label):
        widget.bind("<Button-1>", _set_grouping_disabled)

    # Визуальная подсказка курсором
    col1_wrap.config(cursor="hand2")
    col1_header_label.config(cursor="hand2")
    col1_body_label.config(cursor="hand2")
    col2_wrap.config(cursor="hand2")
    col2_header_label.config(cursor="hand2")
    col2_body_label.config(cursor="hand2")

    def update_grouping_highlight(*_):
        "Обновляет визуальное выделение выбранной колонки по установке флажка."
        def apply_state(frame, header, body, wrap, is_active):
            bg_color = ACTIVE_CARD_COLOR if is_active else FRAME_COLOR
            frame.configure(style="ActiveCard.TFrame" if is_active else "Card.TFrame")
            header.configure(bg=bg_color)
            body.configure(bg=bg_color)
            # Рамка вокруг активной области
            wrap.configure(
                highlightthickness=2 if is_active else 0,
                highlightbackground="#42A5F5" if is_active else FRAME_COLOR,
                highlightcolor="#42A5F5" if is_active else FRAME_COLOR,
                bg=FRAME_COLOR,
            )

        enabled = should_group_by_suffix.get()
        apply_state(col1_frame, col1_header_label, col1_body_label, col1_wrap, enabled)
        apply_state(col2_frame, col2_header_label, col2_body_label, col2_wrap, not enabled)

    should_group_by_suffix.trace_add("write", update_grouping_highlight)
    update_grouping_highlight()

    # --- Кнопка запуска группировки (расположена сразу под колонками) ---
    run_section_frame = ttk.Frame(index_tab_container, padding=(15, 8))
    run_section_frame.pack(side=tk.TOP, fill=tk.X)
    run_card = ttk.Frame(run_section_frame, style="Card.TFrame", padding=15)
    run_card.pack(fill=tk.X, pady=5)

    apply_index_button = ttk.Button(
        run_card,
        text="Применить и группировать",
        command=run_index_packaging,
        style="TButton",
    )
    apply_index_button.pack(fill=tk.X, ipady=10)
    ttk.Label(
        run_card,
        textvariable=index_status_message,
        font=FONT_HELP_TEXT,
        foreground="#757575",
        background=FRAME_COLOR,
        justify=tk.LEFT,
        wraplength=480,
    ).pack(anchor="w", pady=(5, 0))


    # --- Инициализация и привязки ---
    selected_status_key.trace_add("write", update_template_options)
    toggle_delete_option()
    update_template_options()
    # Fallback: если аббревиатура не найдена — выбрать XXX
    if not selected_template_key.get():
        for key, filename in templates_map.items():
            fn_upper = str(filename).upper()
            if "-XXX-" in fn_upper or "(XXX)" in str(key).upper():
                selected_template_key.set(key)
                break

    root.mainloop()

if __name__ == "__main__":
    create_transmittal_gui()
