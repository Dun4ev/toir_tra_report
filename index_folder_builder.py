from __future__ import annotations

import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Callable, TypedDict

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


StatusCallback = Callable[[str], None]

# --- Регулярные выражения для поиска индексов ---
RE_C_GROUPING_KEY = re.compile(
    r"(\b(?:(?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zа-яА-ЯёЁ\-_/])?)(?:-\d{2}-C))\b",
    re.IGNORECASE,
)
RE_GROUPING_KEY = re.compile(
    r"(\b(?:(?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zа-яА-ЯёЁ\-_/])?)(?:-\d{2}-\w{1,2}))\b",
    re.IGNORECASE,
)
RE_INDEX_CODE = re.compile(
    r"(\b(?:[IVXLCDM]+)\.(?:\d+)(?:\.\d+)?(?:\.\d+)?(?:[A-Za-zа-яА-ЯёЁ_/])?)\b",
    re.IGNORECASE,
)

# --- Настройки справочника ---
TZ_SHEET_NAME = "gen_cl"
TZ_LOOKUP_COL = "B"
TZ_PERIODICITY_COL = "E"
TZ_SUFFIX_COL = "G"
TZ_RESERVED_COL = "H"

# --- Таблицы транслитерации ---
CYRILLIC_TO_LATIN = {
    "А": "A",
    "Б": "B",
    "В": "V",
    "Г": "G",
    "Д": "D",
    "Е": "E",
    "Ё": "E",
    "Ж": "ZH",
    "З": "Z",
    "И": "I",
    "Й": "I",
    "К": "K",
    "Л": "L",
    "М": "M",
    "Н": "N",
    "О": "O",
    "П": "P",
    "Р": "R",
    "С": "S",
    "Т": "T",
    "У": "U",
    "Ф": "F",
    "Х": "KH",
    "Ц": "TS",
    "Ч": "CH",
    "Ш": "SH",
    "Щ": "SHCH",
    "Ъ": "",
    "Ы": "Y",
    "Ь": "",
    "Э": "E",
    "Ю": "YU",
    "Я": "YA",
}
CYRILLIC_TO_LATIN.update(
    {k.lower(): v.lower() for k, v in list(CYRILLIC_TO_LATIN.items())}
)
LATIN_TO_CYRILLIC = {
    "a": "а",
    "b": "б",
    "v": "в",
    "g": "г",
    "A": "А",
    "B": "Б",
    "V": "В",
    "G": "Г",
}


PERIODICITY_LATIN_MAP = {
    "М": "M",
    "м": "m",
    "Г": "G",
    "г": "g",
}


class PlannedFolder(TypedDict):
    folder_name: str
    suffix: str | None
    grouping_key: str
    file_paths: list[Path]


class TzSuffixResolver:
    """Загружает файл TZ_glob.xlsx и предоставляет поиск суффиксов для индексов."""

    def __init__(self, tz_file_path: Path) -> None:
        if not tz_file_path.exists():
            raise FileNotFoundError(f"Файл справочника не найден: {tz_file_path}")

        workbook = load_workbook(tz_file_path, data_only=True)
        if TZ_SHEET_NAME not in workbook.sheetnames:
            workbook.close()
            raise ValueError(
                f"На вкладке {tz_file_path.name} отсутствует лист '{TZ_SHEET_NAME}'."
            )

        sheet = workbook[TZ_SHEET_NAME]
        self._lookup_idx = column_index_from_string(TZ_LOOKUP_COL) - 1
        self._periodicity_idx = column_index_from_string(TZ_PERIODICITY_COL) - 1
        self._suffix_idx = column_index_from_string(TZ_SUFFIX_COL) - 1
        self._reserved_idx = column_index_from_string(TZ_RESERVED_COL) - 1
        self._rows = tuple(sheet.iter_rows(values_only=True))
        workbook.close()

    @staticmethod
    def _normalize_lookup(value: object | None) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    @staticmethod
    def _normalize_reserved(value: object | None) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        return text.zfill(2) if text.isdigit() else text.upper()

    @staticmethod
    def _normalize_suffix(value: object | None) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    @staticmethod
    def _normalize_periodicity(value: object | None) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        if not text:
            return None
        transliterated = "".join(PERIODICITY_LATIN_MAP.get(ch, ch) for ch in text)
        return transliterated.upper()

    def find_suffix(
        self,
        lookup_key: str,
        reserved_code: str | None = None,
        periodicity_code: str | None = None,
    ) -> str | None:
        """Возвращает подходящий суффикс по индексу и коду Reserved."""
        normalized_lookup = lookup_key.strip()
        if not normalized_lookup:
            return None

        lookup_variants = {normalized_lookup.lower()}
        last_char = normalized_lookup[-1]
        if last_char in LATIN_TO_CYRILLIC:
            variant = normalized_lookup[:-1] + LATIN_TO_CYRILLIC[last_char]
            lookup_variants.add(variant.lower())

        normalized_reserved = self._normalize_reserved(reserved_code)
        normalized_periodicity = self._normalize_periodicity(periodicity_code)
        fallback: str | None = None

        for row in self._rows:
            if len(row) <= max(
                self._lookup_idx, self._suffix_idx, self._periodicity_idx
            ):
                continue

            row_lookup_raw = (
                row[self._lookup_idx] if len(row) > self._lookup_idx else None
            )
            row_lookup = self._normalize_lookup(row_lookup_raw)
            if row_lookup not in lookup_variants:
                continue

            row_periodicity_raw = (
                row[self._periodicity_idx] if len(row) > self._periodicity_idx else None
            )
            row_periodicity = self._normalize_periodicity(row_periodicity_raw)

            if normalized_periodicity:
                if row_periodicity != normalized_periodicity:
                    continue
            else:
                if row_periodicity:
                    continue

            row_suffix_raw = (
                row[self._suffix_idx] if len(row) > self._suffix_idx else None
            )
            suffix = self._normalize_suffix(row_suffix_raw)
            if not suffix:
                continue

            row_reserved_raw = (
                row[self._reserved_idx] if len(row) > self._reserved_idx else None
            )
            row_reserved = self._normalize_reserved(row_reserved_raw)

            if normalized_reserved:
                if row_reserved == normalized_reserved:
                    return suffix
                if fallback is None:
                    fallback = suffix
            else:
                return suffix

        return fallback


def transliterate_cyrillic_to_latin(text: str) -> str:
    """Преобразует строку, заменяя кириллицу на латиницу для имён директорий."""
    return "".join(CYRILLIC_TO_LATIN.get(ch, ch) for ch in text)


def extract_reserved_value(grouping_key: str) -> str | None:
    """Извлекает значение Reserved из ключа группы вида `II.1.4-02-C`."""
    parts = grouping_key.split("-")
    if len(parts) < 2:
        return None

    candidate = parts[1].strip()
    if not candidate:
        return None

    return candidate.zfill(2) if candidate.isdigit() else candidate.upper()


def extract_periodicity_value(grouping_key: str) -> str | None:
    """Возвращает значение периодичности из ключа группы (`...-XX`)."""
    parts = grouping_key.split("-")
    if len(parts) < 3:
        return None
    value = parts[2].strip()
    return value or None


def _group_files(source_dir: Path) -> dict[str, list[Path]]:
    grouped: dict[str, list[Path]] = defaultdict(list)
    for file_path in source_dir.rglob("*"):
        if not file_path.is_file():
            continue

        c_match = RE_C_GROUPING_KEY.search(file_path.name)
        if c_match:
            grouped[c_match.group(1)].append(file_path)
            continue

        match = RE_GROUPING_KEY.search(file_path.name)
        if match:
            grouped[match.group(1)].append(file_path)

    return grouped


def _transfer_file(src: Path, dest_dir: Path, is_copy: bool) -> Path:
    operation = shutil.copy if is_copy else shutil.move
    dest = dest_dir / src.name
    if not dest.exists():
        transferred = operation(str(src), str(dest))
        return Path(transferred)

    stem = dest.stem
    suffix = dest.suffix
    counter = 1
    while True:
        candidate = dest_dir / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            transferred = operation(str(src), str(candidate))
            return Path(transferred)
        counter += 1


def _notify(callback: StatusCallback | None, message: str) -> None:
    if callback is not None:
        callback(message)


def prepare_index_folders(
    source_dir: Path,
    destination_dir: Path,
    tz_file_path: Path,
    status_callback: StatusCallback | None = None,
    use_copy: bool = True,
    group_by_suffix: bool = False,
) -> list[Path]:
    """Группирует файлы по индексам и перемещает их в целевые каталоги."""
    source_dir = source_dir.resolve()
    destination_dir = destination_dir.resolve()

    if not source_dir.exists() or not source_dir.is_dir():
        raise FileNotFoundError(f"Каталог источника не найден: {source_dir}")

    destination_dir.mkdir(parents=True, exist_ok=True)
    _notify(status_callback, f"Сканирование: {source_dir}")

    files_by_key = _group_files(source_dir)
    if not files_by_key:
        raise ValueError("Не удалось найти файлы с индексами в выбранном каталоге.")

    resolver = TzSuffixResolver(tz_file_path)

    planned_folders: list[PlannedFolder] = []
    items_without_suffix: list[str] = []

    for grouping_key, file_paths in sorted(files_by_key.items()):
        folder_name: str
        suffix: str | None

        if grouping_key.upper().endswith("-C"):
            folder_name = transliterate_cyrillic_to_latin(grouping_key)
            suffix = None  # У C-групп нет суффикса
        else:
            index_match = RE_INDEX_CODE.search(grouping_key)
            if not index_match:
                _notify(
                    status_callback,
                    f"Пропуск: не удалось выделить индекс из {grouping_key}.",
                )
                continue

            index_code = index_match.group(1)
            reserved_code = extract_reserved_value(grouping_key)
            periodicity_code = extract_periodicity_value(grouping_key)
            suffix = resolver.find_suffix(index_code, reserved_code, periodicity_code)

            if not suffix:
                items_without_suffix.append(
                    f"{index_code} (Reserved={reserved_code or '—'}; Periodicity={periodicity_code or '—'})"
                )
                if not group_by_suffix:
                    _notify(
                        status_callback,
                        "Нет суффикса для "
                        f"{index_code} (Reserved={reserved_code or '—'}; Periodicity={periodicity_code or '—'}).",
                    )
                    continue

            latin_key = transliterate_cyrillic_to_latin(grouping_key)
            folder_name = f"{latin_key}_{suffix}" if suffix else latin_key

        planned_folders.append(
            {
                "folder_name": folder_name,
                "suffix": suffix,
                "grouping_key": grouping_key,
                "file_paths": file_paths,
            }
        )

    if group_by_suffix and items_without_suffix:
        error_msg = (
            "Невозможно сгруппировать по суффиксу. Не найдены суффиксы для следующих групп:\n\n"
            + "\n".join(items_without_suffix)
        )
        raise ValueError(error_msg)

    created_dirs: list[Path] = []
    for plan in planned_folders:
        base_dir = destination_dir
        # Если группировка включена и суффикс есть, создаем подпапку
        suffix_value = plan["suffix"]
        if group_by_suffix and suffix_value:
            base_dir = destination_dir / suffix_value

        target_dir = base_dir / plan["folder_name"]
        target_dir.mkdir(parents=True, exist_ok=True)
        if target_dir not in created_dirs:
            created_dirs.append(target_dir)

        _notify(
            status_callback,
            f"Группа {plan['grouping_key']} → {target_dir.relative_to(destination_dir)}",
        )

        for file_path in plan["file_paths"]:
            transferred_path = _transfer_file(file_path, target_dir, is_copy=use_copy)
            _notify(status_callback, f"  • {file_path.name} → {transferred_path.name}")

    _notify(status_callback, "Группировка завершена.")
    return created_dirs


__all__ = [
    "prepare_index_folders",
    "transliterate_cyrillic_to_latin",
    "extract_reserved_value",
    "extract_periodicity_value",
    "TzSuffixResolver",
]
