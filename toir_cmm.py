import re
from pathlib import Path
from datetime import datetime
import sys
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        try:
            print(text.encode('utf-8'))
        except Exception:
            pass # Игнорируем ошибки вывода, чтобы программа не падала

# === НАСТРОЙКИ ===
# --- Определение путей для .exe и обычного режима ---
def get_base_path() -> Path:
    """Возвращает базовый путь для ресурсов, работающий и для .exe."""
    if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
        return Path(sys._MEIPASS)
    else:
        return Path(__file__).parent

BASE_DIR = get_base_path()
TEMPLATE_PATH = BASE_DIR / "Template/CommentSheet_Template.xltx"
TZ_FILE_PATH = BASE_DIR / "Template/TZ_glob.xlsx"
DATE_FMT = "dd.mm.yyyy"

# --- Регулярные выражения (из toir_tra_report_v1.py) ---
RE_INDEX = re.compile(
    r"\b([IVXLCDM]+)\.(\d+)(?:\.(\d+))?(?:\.(\d+))?([A-Za-zА-Яа-я])?\b",
    re.IGNORECASE
)

# --- НОВЫЕ ФУНКЦИИ (из toir_tra_report_v1.py) ---

def normalize_key(key: str) -> str:
    """Нормализует ключ для поиска в словаре."""
    key = key.upper()
    replacements = {'A': 'А', 'B': 'Б', 'V': 'В', 'G': 'Г'}
    for lat, cyr in replacements.items():
        key = key.replace(lat, cyr)
    return key

def extract_index_from_name(filename: str) -> str | None:
    """Извлекает структурированный индекс из имени файла."""
    m = RE_INDEX.search(filename)
    if not m: return None
    roman, num1, num2, num3, suf = m.groups()
    suf = suf or ""
    idx = f"{roman.upper()}.{num1}"
    if num2: idx += f".{num2}"
    if num3: idx += f".{num3}"
    idx += suf
    return idx

def build_tz_map_from_xlsx(xlsx_path: Path) -> dict[str, str]:
    """
    Строит карту {индекс: описание} по всем листам указанного XLSX файла.
    Логика полностью взята из toir_tra_report_v1.py.
    """
    tz_map: dict[str, str] = {}
    if not xlsx_path.exists():
        safe_print(f"  - [WARNING] Файл с данными не найден: {xlsx_path}")
        return tz_map
    
    safe_print(f"  - [INFO] Чтение карты индексов из: {xlsx_path.name}")
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        for ws in wb.worksheets:
            max_col = min(ws.max_column, 20)
            for r in range(1, ws.max_row + 1):
                idx_val, idx_col = None, None
                for c in range(1, max_col + 1):
                    v = ws.cell(r, c).value
                    if isinstance(v, str):
                        m = RE_INDEX.search(v)
                        if m:
                            roman, num1, num2, num3, suf = m.groups()
                            suf = suf or ""
                            idx_val = f"{roman.upper()}.{num1}"
                            if num2: idx_val += f".{num2}"
                            if num3: idx_val += f".{num3}"
                            idx_val += suf
                            idx_col = c
                            break
                if not idx_val: continue
                naziv = None
                vC = ws.cell(r, 3).value
                if isinstance(vC, str) and vC.strip():
                    naziv = vC.strip()
                else:
                    for c in range((idx_col or 1) + 1, max_col + 1):
                        v = ws.cell(r, c).value
                        if isinstance(v, str) and len(v.strip()) >= 3:
                            naziv = v.strip()
                            break
                if naziv:
                    normalized_key = normalize_key(idx_val)
                    if normalized_key not in tz_map:
                        tz_map[normalized_key] = naziv
        safe_print(f"  - [INFO] Карта индексов успешно построена. Найдено {len(tz_map)} записей.")
        return tz_map
    except Exception as e:
        safe_print(f"  - [ERROR] Ошибка при чтении файла {xlsx_path}: {e}")
        return tz_map

# --- СУЩЕСТВУЮЩИЕ ФУНКЦИИ (с изменениями) ---

def ensure_named_range(ws, wb, cell, name):
    """Создаёт именованный диапазон, если ещё не существует."""
    existing = set(wb.defined_names.keys())
    if name not in existing:
        dn = DefinedName(name=name, attr_text=f"'{ws.title}'!{cell.coordinate}")
        wb.defined_names.append(dn)

def fill_basic_fields(wb, report_name: str):
    """Заполнить D1/D4 через именованные диапазоны (если их нет — пишем прямо)."""
    ws = wb.active
    dn_map = dict(wb.defined_names.items())

    # ReportName -> D1
    if "ReportName" in dn_map:
        dests = dn_map["ReportName"].destinations
        for sheet, coord in dests:
            ws_target = wb[sheet] if isinstance(sheet, str) else sheet
            ws_target[coord].value = report_name
    else:
        ws["D1"].value = report_name
        ensure_named_range(ws, wb, ws["D1"], "ReportName")

    # CreatedDate -> D4
    today = datetime.now()
    if "CreatedDate" in dn_map:
        for sheet, coord in dn_map["CreatedDate"].destinations:
            ws_target = wb[sheet] if isinstance(sheet, str) else sheet
            cell = ws_target[coord]
            cell.value = today
            cell.number_format = DATE_FMT
    else:
        ws["D4"].value = today
        ws["D4"].number_format = DATE_FMT
        ensure_named_range(ws, wb, ws["D4"], "CreatedDate")

def fill_extra_fields(wb, report_name: str, tz_map: dict):
    """
    Извлечь код из имени файла, найти его в карте описаний (tz_map)
    и записать результат в ячейку ExtraField1 (D6).
    """
    ws = wb.active
    icode = extract_index_from_name(report_name)

    extra_value = "Код не найден в имени файла"
    if icode:
        safe_print(f"  Найден код в имени файла: {icode}")
        normalized_icode = normalize_key(icode)
        description = tz_map.get(normalized_icode)
        
        if description:
            safe_print(f"  Найдено описание в карте: \"{description}\"")
            extra_value = description
        else:
            safe_print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Код '{icode}' не найден в карте описаний.")
            extra_value = f"ОПИСАНИЕ ДЛЯ {icode} НЕ НАЙДЕНО"
    else:
        safe_print(f"  - [ПРЕДУПРЕЖДЕНИЕ] Код раздела не найден в имени файла: {report_name}")

    # Вставляем найденное значение в ячейку
    dn_map = dict(wb.defined_names.items())
    if "ExtraField1" in dn_map:
        for sheet, coord in dn_map["ExtraField1"].destinations:
            ws_target = wb[sheet] if isinstance(sheet, str) else sheet
            ws_target[coord].value = extra_value
    else:
        ws["D6"].value = extra_value

def make_cmm_for_report(report_path: Path, tz_map: dict):
    """Создает CMM файл для одного отчета."""
    stem = report_path.stem
    cmm_name = f"{stem}_CMM.xlsx"
    cmm_path = report_path.with_name(cmm_name)

    if cmm_path.exists():
        safe_print(f"[ПРОПУСК] Файл уже существует: {cmm_path.name}")
        return

    safe_print(f"Обработка: {report_path.name}")
    try:
        wb = load_workbook(TEMPLATE_PATH)
        wb.template = False
        fill_basic_fields(wb, stem)
        fill_extra_fields(wb, stem, tz_map)
        wb.save(cmm_path)
        safe_print(f"[OK] Создан файл: {cmm_path.name}")
    except Exception as e:
        safe_print(f"[ОШИБКА] Не удалось обработать {report_path.name}: {e}")

def main():
    """Главная функция для пакетной обработки."""
    root = tk.Tk()
    root.withdraw()

    # messagebox.showinfo("Начало работы", "Сейчас вам нужно будет выбрать папку с файлами отчетов (.docx, .pdf)...")
    
    search_dir = filedialog.askdirectory(
        title="Выберите папку с файлами отчетов (.docx, .pdf)"
    )

    if not search_dir:
        messagebox.showwarning("Отмена", "Папка не выбрана. Завершение работы.")
        return
        
    search_path = Path(search_dir)

    # Строим карту ОДИН РАЗ перед началом обработки
    tz_map = build_tz_map_from_xlsx(TZ_FILE_PATH)

    safe_print(f"Запуск пакетной обработки в директории: {search_path.resolve()}")
    
    docx_files = list(search_path.glob("**/*.docx"))
    pdf_files = list(search_path.glob("**/*.pdf"))
    files_to_process = docx_files + pdf_files

    if not files_to_process:
        messagebox.showinfo("Файлы не найдены", f"В директории '{search_path}' и ее подпапках не найдены файлы .docx или .pdf.")
        return

    processed_files = 0
    for doc_file in files_to_process:
        if doc_file.name.startswith("CT-DR-"):
            make_cmm_for_report(doc_file, tz_map)
            processed_files += 1
            
    final_message = f"Обработка завершена.\nВсего найдено файлов (.docx, .pdf): {len(files_to_process)}.\nОбработано (с префиксом CT-DR-): {processed_files}."
    safe_print(final_message.replace('\n', ' ')) # Логируем в одну строку
    messagebox.showinfo("Готово!", final_message)

if __name__ == "__main__":
    main()